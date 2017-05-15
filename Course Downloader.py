# -*- coding: utf-8 -*-
"""
Created on Wed Mar 29 23:07:54 2017

@author: amylz
"""
import os,time,urllib
from selenium import webdriver as webdriver
from selenium.common.exceptions import NoSuchElementException as NoSuchElementException
from urllib import request as urlRequest 
from urllib.request import splittype as splittype
from urllib.request import urlopen as urlopen
from urllib.request import contextlib as contextlib

import openpyxl
from openpyxl import load_workbook  
from openpyxl import Workbook
#from bs4 import BeautifulSoup

"""
user_agent = ("Mozilla/5.0 (Windows NT 10.0; WOW64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/56.0.2924.87 Safari/537.36")

def randomString(length):
    return ''.join(random.choice(string.ascii_letters + string.digits) for i in range(length))
XCSRF2Cookie = 'csrf2_token_%s' % ''.join(randomString(8))
XCSRF2Token = ''.join(randomString(24))
XCSRFToken = ''.join(randomString(24))
cookie = "csrftoken=%s; %s=%s" % (XCSRFToken, XCSRF2Cookie, XCSRF2Token)


headers = {
   # "Accept-Encoding":"gzip, deflate, br",
  #  "Accept-Language":"zh-CN,zh;q=0.8",
  #  "Connection":"keep-alive",
    "User-Agent": user_agent,
    "Host":"accounts.coursera.org",
    "Origin":"https://accounts.coursera.org",
    "Referer": "https://accounts.coursera.org/signin",
    "X-CSRF2-Cookie": XCSRF2Cookie,
    "X-CSRF2-Token": XCSRF2Token,
    "X-CSRFToken": XCSRFToken,
    "Cookie": cookie,
  #  "X-Requested-With":"XMLHttpRequest"
    }
    
coursera_session = requests.session()
login_url = "https://accounts.coursera.org/api/v1/login"

def login(email,password):
    postdata = {
                "email":email,
                "password":password,
                "webrequest":"true"
                }
    login_res = coursera_session.post(login_url, data=postdata, headers=headers)
   # print(login_res.status_code)
    if login_res.status_code == 200:
        print ("Login Successfully!")
        
"""     

# Replace the original urlretrieve to avoid download failure with unstable Internet.
# This function will try several times(depending on argument maxtries) if downloaded file is not complete.
def safe_retrieve(url, filename=None, reporthook=None, data=None, maxtries=5, r_range=None):
    if maxtries < -1:
       raise ValueError('maxtries must be at least equal with -1')
    url_type, path = splittype(url)

           
    with contextlib.closing(urlopen(url, data)) as fp:
        headers = fp.info()
        if not r_range is None:
            try:
                headers["Range"] = "bytes=%d-%d" % r_range
            except TypeError:
                raise ValueError('r_range argument must be a tuple of two int : (start, end)')

        # Just return the local path and the "headers" for file://
        # URLs. No sense in performing a copy unless requested.
        if url_type == "file" and not filename:
            return os.path.normpath(path), headers

        # Handle temporary file setup.
        if filename:
            tfp = open(filename, 'wb')
        else:
            tfp = tempfile.NamedTemporaryFile(delete=False)
            filename = tfp.name
            _url_tempfiles.append(filename)

        with tfp:
            result = filename, headers
            bs = 1024*8
            size = -1
            read = 0
            blocknum = 0
            if "content-length" in headers:
                size = int(headers["Content-Length"])
            elif r_range is not None:
                size = r_range[1]

            if reporthook:
                reporthook(blocknum, bs, size)

            while True:
                block = fp.read(bs)
                if not block:
                    break
                read += len(block)
                tfp.write(block)
                blocknum += 1
                if reporthook:
                    reporthook(blocknum, bs, size)

    if size >= 0 and read < size:
        if maxtries > 0 or maxtries == -1:
           safe_retrieve(url, filename, reporthook, data, maxtries if maxtries == -1 else maxtries-1,r_range=(read, size))
        else:
            raise ContentTooShortError("retrieval incomplete: got only %i out of %i bytes"% (read, size), result)

    return result

    
def get_length(subtitle,video_time):
    for i in range(len(subtitle)-1,-1,-1):
        if subtitle[i] == ':' and subtitle[i-1].isdigit() == True:
            end_minute = int(subtitle[i-1]) + int(subtitle[i-2])*10
            end_second = int(subtitle[i+1])*10 + int(subtitle[i+2])
            break
    for i in range(0,len(subtitle)):
        if subtitle[i] == ':' and subtitle[i-1].isdigit() == True:
            start_minute = int(subtitle[i+1])*10 + int(subtitle[i+2])
            start_second = int(subtitle[i+4])*10 + int(subtitle[i+5])
            break
    video_time[0] += end_minute - start_minute
    video_time[1] += end_second - start_second
        
    
# For Mainland Internet users, it's usually difficult to download coursera videos.  
# If you have applied for any proxy server(I myself used Lantern proxy), set here to boost the downloading process.
proxy_support = urlRequest.ProxyHandler({"https":"..."}) # your proxy port and user ID(if needed)
opener = urlRequest.build_opener(proxy_support)
urlRequest.install_opener(opener)



email = "..." #your coursera ID email
password = "..."    #your password
wait_time = 5  # wait some time so as not to be recognized as machine 

driver = webdriver.Chrome()   #start webdriver

#login first
driver.get('https://www.coursera.org/?authMode=login')   
time.sleep(wait_time)
driver.maximize_window()

video_xpath = "//*[@id=\"rendered-content\"]/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div[2]/ul/span/li/a"
#email_xpath = "//*[@id=\"rendered-content\"]/div/div/div/div/div[3]/div/div/div/form/div[1]/div[1]/input"
#password_xpath = "//*[@id=\"rendered-content\"]/div/div/div/div/div[3]/div/div/div/form/div[1]/div[2]/input"
#submit_xpath = "//*[@id=\"rendered-content\"]/div/div/div/div/div[3]/div/div/div/form/button"
subtitle_xpath = "//*[@id=\"rendered-content\"]/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div/span"

email_field = driver.find_element_by_name("email")
password_field = driver.find_element_by_name("password")
login_button = driver.find_element_by_xpath("//*[@id=\"rendered-content\"]/div/div/div[1]/div[3]/div[2]/div/div[3]/div/div/div/form/button")


#fill in the email andcou password on the webpage to login
email_field.clear()
password_field.clear()
email_field.send_keys(email)   
password_field.send_keys(password)
login_button.click()

course_urls =[]
courses = driver.find_elements_by_class_name('link-button')
for course in courses:
    href = course.get_attribute('href')
    if "learn" in str(href):
        course_urls.append(href)
print("There are " + str(len(course_urls)) + " courses found")

courses_loaded = 63
courses_limit = 23  
 
for i in range(0,courses_limit+1):
    course_url = course_urls[i]  # i = 0,1,2,3...
    driver.get(course_url)
    weeks = driver.find_elements_by_class_name('drawer-item')
    course_title = course_url[course_url.find('learn')+6:course_url.find('/home')]
    start_week = 1
    end_week = len(weeks)
    
    time.sleep(wait_time*2)
    ch_title = driver.find_element_by_class_name("course-name").text
    school = driver.find_element_by_class_name("partner-names").text[4:]

	# estimate the length of video downloaded(optional) 
    workbook = load_workbook('D:/coursera/coursera.xlsx')
    worksheet = workbook.active
    worksheet.cell('A'+str(i+2+courses_loaded)).value = ch_title
    worksheet.cell('C'+str(i+2+courses_loaded)).value = course_title
    worksheet.cell('E'+str(i+2+courses_loaded)).value = school
    video_count = 0
    video_time = [0,0] 

    newpath = "D:/coursera/" + course_title 
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        
    for week in range (start_week, end_week + 1):
        driver.get(course_url+'/week/'+str(week)+'/') 
        #  driver.get(course_url+'/home/week/'+str(week))
        time.sleep(wait_time*2)
        #find the course video elements
        main_collections = driver.find_elements_by_class_name('rc-ItemLink')
        video_urls = []

#process the links of course videos
        for vid in main_collections:
            href = vid.get_attribute('href')
            if "lecture" in str(href):
                video_urls.append(href)
        print("There are " + str(len(video_urls)) + " videos found in week"+str(week))
        video_count += len(video_urls)
        
        for video_index in range(0, len(video_urls)):
            flag = True
            ch_flag = True
            en_flag = True
            driver.get(video_urls[video_index])
            time.sleep(wait_time*2)
        #download_video = driver.find_element_by_xpath(video_xpath)
        #download_video.click();   #Download the video to your own computer
            try:
                vid_link = driver.find_element_by_css_selector('.rc-LectureDownloadItem > a:nth-child(1)').get_attribute('href')
        #download the video to your own path. file name can be changed.
                safe_retrieve(vid_link, newpath + "/" + course_title + '_week' + str(week) + '_' + str(video_index) + ".mp4")
            except NoSuchElementException: 
                flag = False
                print("Video Location Error: " + course_title + '_week' + str(week) + '_' + str(video_index)+".mp4")
        #subtitles = driver.find_elements_by_class_name('rc-Phrase') 
            if(flag == True): 
                try:    
                    for sub_index in range(1,10):
                        sub = driver.find_element_by_xpath("//*[@id=\"c-video_html5_api\"]/track["+str(sub_index)+"]")
                        if(sub.get_attribute('srclang') == 'zh-CN'):
                            break       
                    sub_link = sub.get_attribute('src')
                    subtitle = urlRequest.urlopen(sub_link).read().decode("UTF-8","ignore")
                    #remove spaces
                    sub_str = "".join(subtitle.split(' ')) 
                    get_length(sub_str,video_time)
                except NoSuchElementException:
                    ch_flag = False
                    print("Subtitle Location Error: " + course_title + '_week' + str(week) + '_' + str(video_index)+".srt")
               #download English subtitles
                try:
                    for sub_index in range(1,10):
                        sub_en = driver.find_element_by_xpath("//*[@id=\"c-video_html5_api\"]/track["+str(sub_index)+"]")
                        if(sub_en.get_attribute('srclang') == 'en'):
                            break
                    en_link = sub_en.get_attribute('src')
                    subtitle_en = urlRequest.urlopen(en_link).read().decode("UTF-8","ignore")
                except NoSuchElementException:
                    en_flag = False
                    print("Subtitle Location Error: " + course_title + '_week' + str(week) + '_' + str(video_index)+"_en.srt")                   
            if(flag == True and ch_flag == True):
                try:
                    f = open(newpath + "/" + course_title + '_week' + str(week) + '_' + str(video_index)+".srt","w")
                    f.write(subtitle[6:])
                    f.close()               
                except UnicodeEncodeError:
                    f = open(newpath + "/" + course_title + '_week' + str(week) + '_' + str(video_index)+".srt","w",encoding = "UTF-8")
                    print("Unicode Error: " + course_title + '_week' + str(week) + '_' + str(video_index)+".srt")
                    f.write(subtitle[6:])
                    f.close()
            if(flag == True and en_flag == True):
                try:
                    f_en = open(newpath + "/" + course_title + '_week' + str(week) + '_' + str(video_index)+"_en.srt","w")
                    f_en.write(subtitle_en[6:])
                    f_en.close()               
                except UnicodeEncodeError:
                    f_en = open(newpath + "/" + course_title + '_week' + str(week) + '_' + str(video_index)+"_en.srt","w",encoding = "UTF-8")
                    print("Unicode Error: " + course_title + '_week' + str(week) + '_' + str(video_index)+"_en.srt")
                    f_en.write(subtitle_en[6:])
                    f_en.close()
        
        worksheet.cell('G'+str(i+2+courses_loaded)).value = video_count
        worksheet.cell('H'+str(i+2+courses_loaded)).value = (video_time[0]+video_time[1]/60)/60
        workbook.save('D:/coursera/coursera.xlsx')
     
                    
            
     
    # subtitle_button = driver.find_element_by_class_name("cif-captions")
    # subtitle_button.click()
    # driver.refresh()

     
            #get the subtitle text
            #for subtitle_index in range(len(subtitles)):
                #   subtitle_texts.append(subtitles[subtitle_index].text)
                #   print(subtitle_texts)

driver.quit()
